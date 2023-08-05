import json
import time
import re
from openpyxl import load_workbook
import win32api,win32con
import os

loadfolder = r'plansfolder'
recordfolder = r'record'

class FilesClass:
    def __init__(self):
        self.contect_dict = dict()
        if not os.path.isdir(loadfolder): os.makedirs(loadfolder)
        self.refuse()

    def refuse(self):
        path_list = list()
        for fpath,nousevalue,names in os.walk(loadfolder):
            for name in names:
                pat = os.path.join(fpath,name)
                if os.path.splitext(name)[-1] == '.json':
                    with open(pat,'r+',encoding='utf-8') as fi:
                        self.contect_dict.update({os.path.join(fpath,name):json.loads(fi.read())})
                path_list.append(pat)

        for i in self.contect_dict:
            if i not in path_list:
                del self.contect_dict[i]
        return None

    def getByWeek(self,strj:str='') -> dict:
        alist = dict()
        for i , k in self.contect_dict.items():
            if "week" in k and k["week"] == strj:
                alist.update({i:k})
        return alist

    def getByTime(self,strj:str='') -> dict:
        alist = dict()
        if re.search(r"\d{2}:\d{2}",strj) is None: restr = ''
        else: restr = re.search(r"\d{2}:\d{2}",strj).group()
        for i , k in self.contect_dict.items():
            if "time" in k and k["time"] == restr:
                alist.update({i:k})
        return alist

    def getByEntercode(self,strj:str | int=None) -> dict:
        if strj is None:
            strj = ''
        else:
            if strj.isnumeric():
                strj = int(strj)
            else:pass
        alist = dict()
        for i , k in self.contect_dict.items():
            if "entercode" in k and k["entercode"] == strj:
                alist.update({i:k})
        return alist

    def write(self,index:str='',week:str='',time:str='',entercode:str|int=None):
        if re.search(r"\d{2}:\d{2}",time) is None: time = ''
        else: time = re.search(r"\d{2}:\d{2}",time).group()
        if index != '' or index is not None:
            if index not in self.contect_dict:
                self.contect_dict.update({index:{}})
            if week != '' or week is not None:
                self.contect_dict[index].update({"week":week})
            if time != '' or time is not None:
                self.contect_dict[index].update({"time":time})
            if entercode != '' or entercode is not None:
                self.contect_dict[index].update({"entercode":entercode})

    def save(self):
        for i , k in self.contect_dict.items():
            with open(i,'w+',encoding='utf-8') as fi:
                fi.write(json.dumps(k))


def getRecord(name:str) -> list:
    ReList = []
    if not os.path.isdir(recordfolder): os.makedirs(recordfolder)
    with open(os.path.join(recordfolder,name),'a+',encoding='utf-8') as record:
        record.seek(0,0)
        recordList = record.readlines()
        for cord in recordList:
            if cord != "" or not cord.isspace():
                cord = cord.strip()
                ReList.append(cord)
    return ReList

def writeRecord(name:str,contect:str):
    with open(os.path.join(recordfolder,name),'a+',encoding='utf-8') as record:
        record.write(contect+'\n')

def plans_get_from_xlsx():
    '''
    load plans from xlsx file   \n
    从xlsx文件加载计划
    '''
    if not os.path.exists(r'plansfolder\xl.xlsx') or not os.path.exists(r'plansfolder\xlsx_key.json'):
        win32api.MessageBox(0,'xlsx加载出错','错误',win32con.MB_OK)
        return None
    LoadDictFromXlsx = dict()
    with open(r'plansfolder\xlsx_key.json','r+',encoding='utf-8') as filecontent:
        xlsxKey = json.loads(filecontent.read())
    openedWorkBook = load_workbook(r'plansfolder\xl.xlsx')
    workshe = openedWorkBook[openedWorkBook.sheetnames[0]]
    for timeCell in workshe["1"]:
        if str(timeCell.value) == time.strftime('%A'):
            timecolumn = timeCell.column
    try: TodayPlansKey = workshe.iter_cols(min_row=2,max_row=workshe.max_row,min_col=timecolumn,max_col=timecolumn)
    except: return None
    TimeXlsxTuple = workshe.iter_cols(min_row=2,max_row=workshe.max_row,max_col=1)
    reTodayPlansKey = list(TodayPlansKey)[0]
    reTimeXlsxTuple = list(TimeXlsxTuple)[0]
    for i in range(len(reTodayPlansKey)):
        if str(reTodayPlansKey[i].value) != 'None' and str(reTodayPlansKey[i].value) != '':
            try:
                keyLoad = xlsxKey[str(reTodayPlansKey[i].value)]
                LoadDictFromXlsx.update({str(reTimeXlsxTuple[i].value):keyLoad})
            except: pass
    return(LoadDictFromXlsx)

def plans_info_get():
    '''get plans contect of files and output today todo plans\n
    从文件夹内的JSON文件加载计划'''
    tojoinmeeting = dict()
    path_list = list()
    if not os.path.isdir(loadfolder): os.makedirs(loadfolder)
    for fpath,nousevalue,names in os.walk(loadfolder):
        for name in names:
            if os.path.splitext(name)[-1] == '.json':
                path_list.append(os.path.join(fpath,name))

    for filecontect in path_list:
        with open(filecontect,'r+',encoding='utf-8') as f:
            jsontext = json.loads(f.read())
        try:
            if jsontext['week'] == time.strftime('%A'):
                tojoinmeeting.update({jsontext['time']:jsontext['entercode']})
        except:
            pass
    return(tojoinmeeting)

def checkListContent(checkindex=None,checkdict :dict=None,checknum=-1):
    '''在字典的一定范围内检查是否有相同的索引\n
    :checkindex: 检查的索引
    :checkdict: 一起检查的字典'''
    try:
        checklist = list(checkdict.keys())
        checkindexnum = checklist.index(checkindex)
        if checklist.index(checkindex) >= abs(checknum):
            if checkdict is not None and checkdict[checklist[checkindexnum+checknum]] == checkdict[checkindex]:
                return True
    except: pass
    return False

def closerTimeCheck(todotime):
    '''check if the time is closer
    检查时间是否与目前接近'''
    hour = int(todotime[0:2])
    min = int(todotime[3:])
    if hour > int(time.strftime('%H')) or hour == int(time.strftime('%H')) and min > int(time.strftime('%M')):
        return True
    else:
        return False

#test    
if __name__ == '__main__':
    print(closerTimeCheck('10:10'))
    print(checkListContent('9',{'1':2,'6':7,'9':6}))
    c = FilesClass()
    c.write("123","123","12:03","123")
    print(c.getByEntercode("123"))
