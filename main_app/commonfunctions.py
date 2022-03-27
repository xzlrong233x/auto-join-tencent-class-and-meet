import json
import time
from openpyxl import load_workbook
import win32api,win32con
import os

def plans_get_from_xlsx():
    '''
    load plans from xlsx file   \n
    从xlsx文件加载计划
    '''
    if not os.path.exists(r'plansfolder\xl.xlsx') or not os.path.exists(r'plansfolder\xlsx_key.json'):
        win32api.MessageBox(0,'xlsx加载出错','错误',win32con.MB_OK)
        return None
    LoadDictFromXlsx = dict()
    with open(r'plansfolder\xlsx_key.json','r+',encoding='utf-8') as filecontent:
        xlsxKey = json.loads(filecontent.read())
    openedWorkBook = load_workbook(r'plansfolder\xl.xlsx')
    workshe = openedWorkBook[openedWorkBook.sheetnames[0]]
    for timeCell in workshe["1"]:
        if str(timeCell.value) == time.strftime('%A'):
            timecolumn = timeCell.column
    try: TodayPlansKey = workshe.iter_cols(min_row=2,max_row=workshe.max_row,min_col=timecolumn,max_col=timecolumn)
    except: return None
    TimeXlsxTuple = workshe.iter_cols(min_row=2,max_row=workshe.max_row,max_col=1)
    reTodayPlansKey = list(TodayPlansKey)[0]
    reTimeXlsxTuple = list(TimeXlsxTuple)[0]
    for i in range(len(reTodayPlansKey)):
        if str(reTodayPlansKey[i].value) != 'None' and str(reTodayPlansKey[i].value) != '':
            try:
                keyLoad = xlsxKey[str(reTodayPlansKey[i].value)]
                LoadDictFromXlsx.update({str(reTimeXlsxTuple[i].value):keyLoad})
            except: pass
    return(LoadDictFromXlsx)

def plans_info_get():
    '''get plans contect of files and output today todo plans\n
    从文件夹内的JSON文件加载计划'''
    loadfolder = r'plansfolder'
    tojoinmeeting = dict()
    path_list = list()
    if not os.path.isdir(loadfolder): os.makedirs(loadfolder)
    for fpath,nousevalue,names in os.walk(loadfolder):
        for name in names:
            if os.path.splitext(name)[-1] == '.json':
                path_list.append(os.path.join(fpath,name))

    for filecontect in path_list:
        with open(filecontect,'r+',encoding='utf-8') as f:
            jsontext = json.loads(f.read())
        try:
            if jsontext['week'] == time.strftime('%A'):
                tojoinmeeting.update({jsontext['time']:jsontext['entercode']})
        except:
            pass
    return(tojoinmeeting)

def checkListContent(checkindex=None,checklist :list=None,checkdict :dict=None,checknum=1):
    '''在列表或字典的一定范围内检查是否有相同的索引,是则返回下一个索引\n
    :checkindex: 检查的索引
    :checklist: 一起检查的列表
    :checkdict: 一起检查的字典'''
    try:
        if checklist.index(checkindex) >= checknum:
            if checkdict is None and checklist[checklist.index(checkindex)-checknum] == checkindex:
                return checklist[checklist.index(checkindex)+checknum]
            elif checkdict[checklist[checklist.index(checkindex)-checknum]] == checkdict[checkindex]:
                return checklist[checklist.index(checkindex)+checknum]
    except IndexError:
        return None
    except:
        pass
    return checkindex

def closerTimeCheck(todotime):
    '''check if the time is closer
    检查时间是否与目前接近'''
    hour = int(todotime[0:2])
    min = int(todotime[3:])
    if hour > int(time.strftime('%H')):
        return True
    elif hour == int(time.strftime('%H')) and min > int(time.strftime('%M')):
        return True
    else:
        return False